#!/usr/bin/env python3
"""
Export unified_statements to Excel with separate tabs by format

Tabs:
- UATL_format_1: Airtel format_1 statements
- UATL_format_2: Airtel format_2 statements
- UMTN_excel: MTN excel statements
- Summary: Aggregated statistics

Output: docs/analysis/unified_statements_export.xlsx
"""

import sys
import os
from pathlib import Path
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Setup paths
load_dotenv(Path(__file__).parent.parent.parent / '.env')
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

def get_database_engine():
    """Create database engine."""
    return create_engine(
        f"mysql+pymysql://{os.getenv('DB_USER')}:{os.getenv('DB_PASSWORD')}@"
        f"{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}",
        pool_pre_ping=True
    )


def get_statements_by_format(engine, acc_prvdr_code, format_type):
    """Query unified_statements for specific provider and format."""
    query = text("""
        SELECT
            u.run_id, u.acc_number, u.acc_prvdr_code, u.format, u.mime,
            u.submitted_date, u.start_date, u.end_date, u.rm_name,
            u.custom_verification, u.custom_verification_reason,
            u.balance_match, u.balance_diff_change_ratio,
            u.stmt_opening_balance, u.stmt_closing_balance,
            u.calculated_closing_balance, u.balance_diff_changes,
            u.credits, u.debits, u.fees, u.charges,
            u.duplicate_count, u.quality_issues_count,
            u.header_row_manipulation_count, u.gap_related_balance_changes,
            u.meta_title, u.meta_author, u.meta_producer,
            u.meta_created_at, u.meta_modified_at,
            u.summary_customer_name, u.summary_mobile_number,
            cd.cust_id, cd.borrower_biz_name
        FROM unified_statements u
        LEFT JOIN customer_details cd ON u.run_id = cd.run_id
        WHERE u.acc_prvdr_code = :acc_prvdr_code
          AND u.format = :format_type
        ORDER BY u.submitted_date DESC
    """)

    with engine.connect() as conn:
        df = pd.read_sql(query, conn, params={
            'acc_prvdr_code': acc_prvdr_code,
            'format_type': format_type
        })

    return df


def get_summary_stats(engine):
    """Generate summary statistics."""
    query = text("""
        SELECT
            format,
            acc_prvdr_code,
            custom_verification,
            COUNT(*) as count,
            ROUND(AVG(balance_diff_change_ratio) * 100, 2) as avg_balance_diff_pct,
            SUM(credits) as total_credits,
            SUM(debits) as total_debits
        FROM unified_statements
        WHERE acc_prvdr_code IN ('UATL', 'UMTN')
        GROUP BY format, acc_prvdr_code, custom_verification
        ORDER BY acc_prvdr_code, format,
            CASE custom_verification
                WHEN 'FATAL' THEN 1
                WHEN 'CRITICAL' THEN 2
                WHEN 'NO_ISSUES' THEN 3
                ELSE 4
            END
    """)

    with engine.connect() as conn:
        df = pd.read_sql(query, conn)

    return df


def format_excel_sheet(writer, sheet_name):
    """Apply formatting to Excel sheet."""
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Freeze top row
    worksheet.freeze_panes = 'A2'


def main():
    print("=" * 80)
    print("EXPORT UNIFIED STATEMENTS TO EXCEL")
    print("=" * 80)

    engine = get_database_engine()
    output_dir = Path('docs/analysis')
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / 'unified_statements_export.xlsx'

    # Collect data
    print("\n1. Querying UATL_format_1...")
    uatl_format_1 = get_statements_by_format(engine, 'UATL', 'format_1')
    print(f"   ✓ {len(uatl_format_1):,} records")

    print("\n2. Querying UATL_format_2...")
    uatl_format_2 = get_statements_by_format(engine, 'UATL', 'format_2')
    print(f"   ✓ {len(uatl_format_2):,} records")

    print("\n3. Querying UMTN_excel...")
    umtn_excel = get_statements_by_format(engine, 'UMTN', 'excel')
    print(f"   ✓ {len(umtn_excel):,} records")

    print("\n4. Generating summary...")
    summary = get_summary_stats(engine)
    print(f"   ✓ {len(summary):,} summary rows")

    # Write to Excel
    print(f"\n5. Writing to Excel: {output_file}")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write data tabs
        uatl_format_1.to_excel(writer, sheet_name='UATL_format_1', index=False)
        format_excel_sheet(writer, 'UATL_format_1')

        uatl_format_2.to_excel(writer, sheet_name='UATL_format_2', index=False)
        format_excel_sheet(writer, 'UATL_format_2')

        umtn_excel.to_excel(writer, sheet_name='UMTN_excel', index=False)
        format_excel_sheet(writer, 'UMTN_excel')

        summary.to_excel(writer, sheet_name='Summary', index=False)
        format_excel_sheet(writer, 'Summary')

    print(f"   ✓ Excel file created: {output_file}")

    print("\n" + "=" * 80)
    print("EXPORT COMPLETE")
    print("=" * 80)
    print(f"\nOutput file: {output_file}")
    print(f"File size: {output_file.stat().st_size / 1024 / 1024:.2f} MB")

    return 0


if __name__ == '__main__':
    exit(main())
